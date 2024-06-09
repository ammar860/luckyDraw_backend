
import os.path
from database import Database
from configparser import ConfigParser
import openpyxl

config = ConfigParser()
config.read("config.ini")
database = Database(config)


def readExcel():
    conn = database.connect()
    if conn is not None:
        cursor = conn.cursor()
    path = f"{os.getcwd()}\Book2.xlsx"

    print(path)
    wbb = openpyxl.load_workbook(path)

    print("WORK BOOK")
    # sheet_names = wbb.get_sheet_names()
    # sheet_name = sheet_names[0]

    sheet = wbb.active

    max_row = sheet.max_row
    max_col = sheet.max_column

    counter = 0
    dataToSend = []
    for row in sheet.iter_rows(min_row=4, max_col=max_col, max_row=sheet.max_row, values_only=True):
        # counter+=1
        print(row)
        # print("1", row[1])
        # print("2", row[2])
        # print("3", row[3])
        # print("4", row[4])
        # print("5", row[5])
        # print("10", row[10])
        # print("11", row[11])
        # print("12", row[12])
        # chOne = str(row[7])[-1:]
        # chTwo = str(row[8])[-1:]
        # print("CHOICE One", chOne)
        # print("CHOICE Two", chTwo)

        insertDataQuery = f"""INSERT INTO luckyDrawJS(ArmyNo ,Rank, Trade, Name, Arm, JCO_Sldr)
                                            VALUES (?,?,?,?,?,?) """
        cursor.execute(insertDataQuery, (row[1], str(row[2]).strip(), str(row[3]).strip(), str(row[4]).strip(), str(row[5]).strip(), str(row[6]).strip()))
    conn.commit()
    print('ALL DATA INSERTED')


readExcel()